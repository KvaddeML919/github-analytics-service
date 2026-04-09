"""Excel and console output formatting."""

from typing import Any, Dict, List, Optional

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_ALT_ROW_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
_THIN_BORDER = Border(
    bottom=Side(style="thin", color="B4C6E7"),
)

COLUMNS = [
    ("Username",                  "username",                  16),
    ("Total PRs",                 "total_prs",                 10),
    ("PRs / Working Day",         "prs_per_working_day",       16),
    ("Merged PRs",                "merged_prs",                11),
    ("Merge Rate %",              "merge_rate_pct",            12),
    ("Avg Merge Time (hrs)",      "avg_merge_time_hrs",        20),
    ("Total Commits",             "total_commits",             13),
    ("Commits / Day",             "commits_per_coding_day",    14),
    ("Coding Days / Week",        "avg_coding_days_per_week",  18),
    ("Weekend Commits",           "weekend_commits",           16),
    ("Active Repos",              "active_repos",              12),
    ("Reviews Given",             "reviews_given",             14),
    ("PRs Commented On",          "prs_commented_on",          17),
]

TEAM_AVG_KEYS = [
    "prs_per_working_day",
    "merge_rate_pct",
    "avg_merge_time_hrs",
    "commits_per_coding_day",
    "avg_coding_days_per_week",
    "weekend_commits",
    "reviews_given",
    "prs_commented_on",
]

_TEAM_AVG_FONT = Font(bold=True, color="FFFFFF", size=11)
_TEAM_AVG_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

Row = Dict[str, Any]


def compute_team_averages(rows: List[Row]) -> Dict[str, Any]:
    """Return a dict with team-average values for the keys in TEAM_AVG_KEYS.

    Keys not in TEAM_AVG_KEYS are left blank. None values are excluded from
    the average (so a user with None merge time doesn't drag it down).
    """
    if not rows:
        return {}
    avgs: Dict[str, Any] = {"username": "TEAM AVERAGE"}
    for key in TEAM_AVG_KEYS:
        vals = [r[key] for r in rows if r.get(key) is not None]
        if vals:
            avgs[key] = round(sum(vals) / len(vals), 1)
        else:
            avgs[key] = None
    for _, key, _ in COLUMNS:
        if key not in avgs:
            avgs[key] = ""
    return avgs


def write_stats_sheet(ws: Any, rows: List[Row], team_avg: Optional[Row] = None) -> None:
    """Write a formatted stats table into an openpyxl worksheet.

    ``ws`` is an openpyxl Worksheet. ``rows`` is a list of per-user result
    dicts whose keys match the COLUMNS definitions. ``team_avg`` is an
    optional averages row appended below a blank separator line.
    """
    for col_idx, (title, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, r in enumerate(rows, 2):
        for col_idx, (_, key, _) in enumerate(COLUMNS, 1):
            val = r.get(key)
            if val is None:
                val = ""
            ws.cell(row=row_idx, column=col_idx, value=val)

        if row_idx % 2 == 0:
            for col_idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = _ALT_ROW_FILL

        for col_idx in range(1, len(COLUMNS) + 1):
            ws.cell(row=row_idx, column=col_idx).border = _THIN_BORDER

    if team_avg is not None:
        avg_row = len(rows) + 3
        for col_idx, (_, key, _) in enumerate(COLUMNS, 1):
            val = team_avg.get(key)
            if val is None:
                val = ""
            cell = ws.cell(row=avg_row, column=col_idx, value=val)
            cell.font = _TEAM_AVG_FONT
            cell.fill = _TEAM_AVG_FILL
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions


def _fmt_val(val: Any, fmt: str = "", prefix: str = "", suffix: str = "", na: str = "N/A") -> str:
    """Format a value for console display, handling None gracefully."""
    if val is None or val == "":
        return na
    return f"{prefix}{val:{fmt}}{suffix}"


def print_console_tables(results: List[Row], team_avg: Optional[Row] = None) -> None:
    """Print the Activity and Collaboration summary tables to stdout."""
    print(f"\n{'─' * 110}")
    print("ACTIVITY")
    hdr1 = (f"{'Username':<20} {'PRs':>6} {'PRs/Day':>8} {'Merged%':>8} "
            f"{'Commits':>8} {'Commits/Day':>12} {'Coding Days':>12} "
            f"{'Wknd Commits':>13}")
    print(hdr1)
    print("─" * len(hdr1))

    def _print_activity_row(r: Row) -> None:
        cd_str = _fmt_val(r.get("avg_coding_days_per_week"))
        prs = _fmt_val(r.get("total_prs"), na="")
        commits = _fmt_val(r.get("total_commits"), na="")
        wknd = _fmt_val(r.get("weekend_commits"), na="")
        print(f"{r.get('username', ''):<20} "
              f"{prs:>6} "
              f"{_fmt_val(r.get('prs_per_working_day')):>8} "
              f"{_fmt_val(r.get('merge_rate_pct'), suffix='%'):>8} "
              f"{commits:>8} "
              f"{_fmt_val(r.get('commits_per_coding_day')):>12} "
              f"{cd_str:>12} "
              f"{wknd:>13}")

    for r in results:
        _print_activity_row(r)
    if team_avg:
        print("─" * len(hdr1))
        _print_activity_row(team_avg)

    print(f"\n{'─' * 120}")
    print("COLLABORATION & QUALITY")
    hdr2 = (f"{'Username':<20} {'Reviews':>8} {'Commented':>10} "
            f"{'Merge Time':>11} {'Repos':>6}")
    print(hdr2)
    print("─" * len(hdr2))

    def _print_collab_row(r: Row) -> None:
        merge_str = _fmt_val(r.get("avg_merge_time_hrs"), suffix="h")
        reviews = _fmt_val(r.get("reviews_given"), na="")
        commented = _fmt_val(r.get("prs_commented_on"), na="")
        repos = _fmt_val(r.get("active_repos"), na="")
        print(f"{r.get('username', ''):<20} "
              f"{reviews:>8} "
              f"{commented:>10} "
              f"{merge_str:>11} "
              f"{repos:>6}")

    for r in results:
        _print_collab_row(r)
    if team_avg:
        print("─" * len(hdr2))
        _print_collab_row(team_avg)
